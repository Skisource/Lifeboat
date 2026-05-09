"""Excel file parsing for familiarisation and POB lists."""

import re
from datetime import date, datetime
from typing import BinaryIO

import openpyxl


def parse_expiry_date(value: str) -> date | None:
    """
    Parse expiry date from string.

    Args:
        value: Date string in DD/MM/YYYY format or "Required"

    Returns:
        date object if valid date, None if "Required" or invalid
    """
    if not value or not isinstance(value, str):
        return None

    value = value.strip()

    if value.lower() == "required":
        return None

    try:
        return datetime.strptime(value, "%d/%m/%Y").date()
    except ValueError:
        return None


def is_due_for_familiarisation(
    expiry_value: str, reference_date: date | None = None
) -> bool:
    """
    Check if a person is due for familiarisation.

    Args:
        expiry_value: The expiry date string or "Required"
        reference_date: Date to compare against (defaults to today)

    Returns:
        True if person needs familiarisation (Required or date <= reference_date)
    """
    if reference_date is None:
        reference_date = date.today()

    if not expiry_value or not isinstance(expiry_value, str):
        return False

    expiry_value = expiry_value.strip()

    if expiry_value.lower() == "required":
        return True

    expiry_date = parse_expiry_date(expiry_value)
    if expiry_date is None:
        return False

    return expiry_date <= reference_date


def normalize_name(name: str) -> str:
    """
    Normalize a name for comparison.

    Handles both "First Last" and "Last, First" formats.
    Returns lowercase "first last" format.
    """
    name = name.strip()

    if "," in name:
        parts = name.split(",", 1)
        if len(parts) == 2:
            last, first = parts
            name = f"{first.strip()} {last.strip()}"

    name = re.sub(r"\s+", " ", name)
    return name.lower()


def parse_familiarisation_list(
    file: BinaryIO, reference_date: date | None = None
) -> list[str]:
    """
    Parse familiarisation required list from xlsx file.

    Only includes personnel who are due for familiarisation:
    - Expiry date is "Required" (first-timers), OR
    - Expiry date is <= reference_date (overdue or due today)

    Expected format:
    - Row 3: Headers (Site, Full Name, Occupation, First Expiry)
    - Row 4+: Data with names in column B, expiry in column D

    Args:
        file: File-like object containing xlsx data
        reference_date: Date to filter against (defaults to today)

    Returns:
        List of personnel names due for familiarisation

    Raises:
        ValueError: If file format is invalid or required columns missing
    """
    if reference_date is None:
        reference_date = date.today()

    try:
        wb = openpyxl.load_workbook(file, read_only=True, data_only=True)
    except Exception as e:
        raise ValueError(f"Could not read familiarisation file: {e}") from e

    ws = wb.active
    if ws is None:
        raise ValueError("Familiarisation file has no active worksheet")

    names = []
    for row in ws.iter_rows(min_row=4, values_only=True):
        if len(row) < 4:
            continue

        name = row[1]
        expiry = row[3]

        if not name or not isinstance(name, str) or not name.strip():
            continue

        expiry_str = str(expiry) if expiry else ""

        if is_due_for_familiarisation(expiry_str, reference_date):
            names.append(name.strip())

    wb.close()

    if not names:
        raise ValueError("No personnel due for familiarisation found")

    return names


def parse_pob_list(file: BinaryIO) -> list[str]:
    """
    Parse Personnel On Board list from xlsx file.

    Expected format:
    - Row 2: Headers (Person Onboard, Cabin, Occupation, ...)
    - Row 3+: Data with names in column A, but includes category/department rows

    Args:
        file: File-like object containing xlsx data

    Returns:
        List of personnel names currently on board

    Raises:
        ValueError: If file format is invalid or required columns missing
    """
    try:
        wb = openpyxl.load_workbook(file, read_only=True, data_only=True)
    except Exception as e:
        raise ValueError(f"Could not read POB file: {e}") from e

    ws = wb.active
    if ws is None:
        raise ValueError("POB file has no active worksheet")

    names = []
    skip_patterns = [
        r"^Category:",
        r"^Department:",
        r"^Total POB:",
        r"^Person Onboard$",
    ]
    skip_regex = re.compile("|".join(skip_patterns), re.IGNORECASE)

    for row in ws.iter_rows(min_row=3, values_only=True):
        if not row or not row[0]:
            continue

        cell_value = row[0]
        if not isinstance(cell_value, str):
            continue

        cell_value = cell_value.strip()
        if not cell_value:
            continue

        if skip_regex.search(cell_value):
            continue

        if "," in cell_value:
            names.append(cell_value)

    wb.close()

    if not names:
        raise ValueError("No names found in POB file")

    return names


def cross_reference(fam_names: list[str], pob_names: list[str]) -> list[str]:
    """
    Find personnel who are both on familiarisation list and currently on board.

    Handles different name formats:
    - Familiarisation uses "First Last"
    - POB uses "Last, First"

    Args:
        fam_names: Names requiring familiarisation
        pob_names: Names currently on board (in "Last, First" format)

    Returns:
        List of matched personnel names in display format (sorted alphabetically)
    """
    fam_normalized = {normalize_name(name): name for name in fam_names}
    pob_normalized = {normalize_name(name): name for name in pob_names}

    matched_keys = set(fam_normalized.keys()) & set(pob_normalized.keys())

    matched = [pob_normalized[key] for key in matched_keys]
    return sorted(matched, key=lambda x: normalize_name(x))
